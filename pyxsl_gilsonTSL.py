from openpyxl import Workbook, load_workbook
import os
import re
from collections import OrderedDict
from openpyxl.styles import PatternFill

dest_filepath= '/users/spencertrinh/Documents/pythonScripts/pyxsl/'
filenames = []
purpleFill = PatternFill(start_color='CC3366',
                   end_color='CC6699',
                   fill_type='solid')
# sorted_empty_wells = {}

#FOR EACH EXCEL FILE, LOAD THE WORKBOOK AND APPEND TO DICTIONARY WITH WELL_ID AND PLATE_NUM
def make_sorted_empty_well_dict(dest_filepath,start_row_pos):
    # global filenames
    global sorted_empty_wells
    empty_wells = {}
    for i in range(16):
        for j in range(24):
                 empty_wells[f"{chr(65+i)}{j+1}"] =  []
    for f in os.listdir(dest_filepath):
        wb = load_workbook(dest_filepath + f)
        ws = wb.active
        for r in ws.iter_rows(min_row=start_row_pos, values_only=True):
            if re.search('\d+', f):
                empty_wells[r[1]].append(f.split('.')[0].replace('data_',''))
                # plt_code = f.split('.')[0]
                # suffix = plt_code[-3:]
                # plt_code = plt_code[:8]
                # plt_code = plt_code + "_" + suffix
                # empty_wells[r[1]].append(plt_code)
                # filenames.append(plt_code)

    sorted_empty_wells = {k:sorted(v,key=lambda x: int(x)) for k,v in empty_wells.items() if v}


def add_labels():
    #ADD PLATE NUMBERS FOR HEADER
    # header = [f"{nm}" for nm in filenames)]
    header = [f"plate_{i+25}" if i> 25 else f"plate_{i}" for i in range(1,36) ]
    for c in range(len(header)):
        ws.cell(row=1, column=c+2, value=header[c])

    #ADD WELL_IDS ON FIRST COLUMN
    for idx,k in enumerate(sorted_empty_wells.keys()):
        ws.cell(row = idx+2, column = 1).value = k 


#FILL IN CELLS WITH PURPLE COLOUR FROM THE SORTED EMPTY WELL DICTIONARY
def fill_colour_xl():
    for i,(k,v) in enumerate(sorted_empty_wells.items()):
        ws.cell(row = len(sorted_empty_wells.keys())+3+i, column = 2, value = k) # add just text at the bottom
        for j, plt_nums in enumerate(v):
            if int(plt_nums) > 50:
                plt_nums = int(plt_nums)-24
            else:
                plt_nums = int(plt_nums) + 1
                # print(f"{i+2} - {k} - {plt_nums}")
            ws.cell(row = i+2, column = plt_nums).fill = purpleFill
            ws.cell(row = len(sorted_empty_wells.keys())+3+i, column = j+3, value = plt_nums) # add just text


if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    make_sorted_empty_well_dict(dest_filepath,1)
    add_labels()
    fill_colour_xl()
    wb.save(dest_filepath + 'out.xlsx')
